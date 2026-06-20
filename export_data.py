import openpyxl, json
from collections import Counter

wb = openpyxl.load_workbook('earthquake_alert_balanced_dataset(clean)_new.xlsx')
ws = wb.active
headers = [cell.value for cell in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    rows.append(dict(zip(headers, row)))

# Compute validation stats
mags = [r['magnitude'] for r in rows]
depths = [r['depth'] for r in rows]
sigs = [r['sig'] for r in rows]
alerts = [r['alert'] for r in rows]

print('Total:', len(rows))
print('Avg mag:', round(sum(mags)/len(mags), 2))
print('Avg depth:', round(sum(depths)/len(depths), 2))
print('Avg sig:', round(sum(sigs)/len(sigs), 2))

ac = Counter(alerts)
print('Alert counts:', dict(ac))

for level in ['green', 'yellow', 'orange', 'red']:
    subset = [r for r in rows if r['alert'] == level]
    def avg(col, s=subset):
        return round(sum(r[col] for r in s) / len(s), 2)
    mag_avg = round(sum(r['magnitude'] for r in subset) / len(subset), 2)
    dep_avg = round(sum(r['depth'] for r in subset) / len(subset), 2)
    cdi_avg = round(sum(r['cdi'] for r in subset) / len(subset), 2)
    mmi_avg = round(sum(r['mmi'] for r in subset) / len(subset), 2)
    sig_avg = round(sum(r['sig'] for r in subset) / len(subset), 2)
    print(f'{level}: mag={mag_avg} depth={dep_avg} cdi={cdi_avg} mmi={mmi_avg} sig={sig_avg}')

# Export as JS file
json_data = json.dumps(rows, separators=(',', ':'))
print('JSON array length (chars):', len(json_data))
js_content = 'const EARTHQUAKE_DATA = ' + json_data + ';'
with open('eq_data.js', 'w', encoding='utf-8') as f:
    f.write(js_content)
print('Written eq_data.js')
